# ============================================================
# TP5 - TSP : Tournée des capitales et moughataas de Mauritanie
# Recherche Opérationnelle | IMT Mines Alès
# Solveur : OR-Tools (CBC) via ortools.linear_solver
# ============================================================

from ortools.linear_solver import pywraplp
import openpyxl as op
import time


def parse_distance(val):
    """Convertit une valeur Excel en float (gère les virgules et espaces)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(',', '').replace(' ', ''))


def resoudre_tsp(villes, dist, nom_instance):
    """
    Résout le TSP par programmation linéaire en nombres entiers (MTZ).
    Retourne la distance totale et la tournée optimale.
    """
    n = len(villes)

    solver = pywraplp.Solver(f'TSP_{nom_instance}',
                              pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()

    # x[i][j] = 1 si on va de ville i à ville j
    x = [[solver.IntVar(0, 1, f'x_{i}_{j}') for j in range(n)]
         for i in range(n)]

    # u[i] : variable d'ordre pour contraintes MTZ (sous-tournées)
    u = [solver.IntVar(0, n - 1, f'u_{i}') for i in range(n)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Chaque ville visitée exactement une fois (entrée et sortie)
    for i in range(n):
        solver.Add(sum(x[i][j] for j in range(n) if j != i) == 1)  # sortie
        solver.Add(sum(x[j][i] for j in range(n) if j != i) == 1)  # entrée
        x[i][i].SetBounds(0, 0)  # pas de boucle

    # Contraintes MTZ pour éliminer les sous-tournées
    for i in range(1, n):
        for j in range(1, n):
            if i != j:
                solver.Add(u[i] - u[j] + n * x[i][j] <= n - 1)

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    # Minimiser la distance totale
    solver.Minimize(
        sum(dist[i][j] * x[i][j]
            for i in range(n) for j in range(n) if i != j)
    )

    start  = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        distance = solver.Objective().Value()
        print(f"\nDistance totale optimale  : {distance:.0f} km")
        print(f"Temps de calcul           : {elapsed:.4f} s")
        print(f"Itérations                : {solver.iterations()}")
        print(f"Nœuds B&B                 : {solver.nodes()}")

        # Reconstruction de la tournée
        tournee = [0]
        current = 0
        for _ in range(n - 1):
            for j in range(n):
                if j != current and x[current][j].solution_value() > 0.5:
                    tournee.append(j)
                    current = j
                    break
        tournee.append(0)

        print("\nTournée optimale :")
        print(" → ".join(villes[i] for i in tournee))
        return distance, tournee
    else:
        print("Pas de solution optimale trouvée.")
        return None, None


# ─────────────────────────────────────────────
# Exercice 1 — 13 capitales de wilayas
# ─────────────────────────────────────────────
def exercice1_capitales_wilayas():
    print("=" * 55)
    print("Exercice 1 — TSP : 13 capitales de wilayas")
    print("=" * 55)

    file = 'documents/Interface_TP5_Exo1.xlsx'
    wb = op.load_workbook(file)
    ws = wb.active

    n      = 13
    villes = [ws.cell(2 + i, 1).value for i in range(n)]
    dist   = [[parse_distance(ws.cell(2 + i, 2 + j).value)
               for j in range(n)] for i in range(n)]

    distance, tournee = resoudre_tsp(villes, dist, 'Wilayas')

    if distance is not None:
        # Écriture résultats Excel
        try:
            ws_out = wb['Résultats']
        except KeyError:
            ws_out = wb.create_sheet('Résultats')
        ws_out['A1'] = 'Distance totale (km)'
        ws_out['B1'] = round(distance, 0)
        ws_out['A3'] = 'Tournée optimale'
        for k, idx in enumerate(tournee):
            ws_out.cell(4, 1 + k).value = villes[idx]
        wb.save(file)
        print(f"\nRésultats écrits dans : {file}")


# ─────────────────────────────────────────────
# Exercice 2 — 48 moughataas
# ─────────────────────────────────────────────
def exercice2_moughataas():
    print("\n" + "=" * 55)
    print("Exercice 2 — TSP : 48 moughataas de Mauritanie")
    print("=" * 55)

    file = 'documents/Interface_TP5_Exo2.xlsx'
    wb = op.load_workbook(file)
    ws = wb.active

    n      = 48
    villes = [ws.cell(3 + i, 1).value for i in range(n)]
    dist   = [[parse_distance(ws.cell(3 + i, 3 + j).value)
               for j in range(n)] for i in range(n)]

    print(f"⚠️  Problème de taille {n} villes — temps de calcul potentiellement long.")
    distance, tournee = resoudre_tsp(villes, dist, 'Moughataas')

    if distance is not None:
        try:
            ws_out = wb['Résultats']
        except KeyError:
            ws_out = wb.create_sheet('Résultats')
        ws_out['A1'] = 'Distance totale (km)'
        ws_out['B1'] = round(distance, 0)
        ws_out['A3'] = 'Tournée optimale'
        for k, idx in enumerate(tournee):
            ws_out.cell(4, 1 + k).value = villes[idx]
        wb.save(file)
        print(f"\nRésultats écrits dans : {file}")


# ─────────────────────────────────────────────
if __name__ == "__main__":
    exercice1_capitales_wilayas()
    exercice2_moughataas()