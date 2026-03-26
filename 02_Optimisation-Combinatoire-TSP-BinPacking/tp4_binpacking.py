# ============================================================
# TP4 - Bin Packing et Cutting Stock
# Recherche Opérationnelle | IMT Mines Alès
# Solveur : OR-Tools (CBC) via ortools.linear_solver
# ============================================================

from ortools.linear_solver import pywraplp
import openpyxl as op
import time


# ─────────────────────────────────────────────
# Exercice 1 — Chargement avions de fret (Bin Packing)
# ─────────────────────────────────────────────
def exercice1_chargement_avions():
    print("=" * 55)
    print("Exercice 1 — Chargement avions de fret (Bin Packing)")
    print("=" * 55)

    file = 'documents/tp4exo1.xlsx'
    wb = op.load_workbook(file)
    ws = wb.active

    # Lecture des données
    n        = 10       # nombre de types d'objets
    capacity = ws.cell(4, 2).value  # capacité d'un avion

    poids  = [ws.cell(2, 2 + i).value for i in range(n)]
    nombre = [ws.cell(3, 2 + i).value for i in range(n)]
    B      = ws.cell(5, 2).value    # nombre max de vols (bins)

    # Expansion : créer un objet par unité
    items = []
    for i in range(n):
        for _ in range(nombre[i]):
            items.append(poids[i])
    N = len(items)

    solver = pywraplp.Solver('BinPacking_Avions',
                              pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    # x[i][b] = 1 si objet i dans bin b
    x = [[solver.IntVar(0, 1, f'x_{i}_{b}') for b in range(B)]
         for i in range(N)]
    # y[b] = 1 si bin b est utilisé
    y = [solver.IntVar(0, 1, f'y_{b}') for b in range(B)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Chaque objet dans exactement 1 bin
    for i in range(N):
        solver.Add(sum(x[i][b] for b in range(B)) == 1)

    # Capacité de chaque bin
    for b in range(B):
        solver.Add(sum(items[i] * x[i][b] for i in range(N)) <= capacity * y[b])

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    # Minimiser le nombre de bins utilisés
    solver.Minimize(sum(y[b] for b in range(B)))

    start  = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        nb_vols = int(solver.Objective().Value())
        print(f"\nNombre optimal de vols : {nb_vols}")
        print(f"Temps de calcul        : {elapsed:.4f} s")
        print(f"Itérations             : {solver.iterations()}")
        print(f"Nœuds B&B              : {solver.nodes()}")

        # Écriture résultats Excel
        try:
            ws_out = wb['Résultats']
        except KeyError:
            ws_out = wb.create_sheet('Résultats')
        ws_out['A1'] = 'Nombre optimal de vols'
        ws_out['B1'] = nb_vols
        for b in range(B):
            if y[b].solution_value() > 0.5:
                charge = sum(items[i] * x[i][b].solution_value()
                             for i in range(N))
                ws_out.cell(3 + b, 1).value = f'Vol {b+1}'
                ws_out.cell(3 + b, 2).value = round(charge, 1)
        wb.save(file)
        print(f"Résultats écrits dans  : {file}")
    else:
        print("Pas de solution optimale trouvée.")


# ─────────────────────────────────────────────
# Exercice 2 — Découpe bobines (min bobines utilisées)
# ─────────────────────────────────────────────
def exercice2_cutting_stock_min_bobines():
    print("\n" + "=" * 55)
    print("Exercice 2 — Découpe bobines (min bobines utilisées)")
    print("=" * 55)

    file = 'documents/tp4exo3.xlsx'
    wb   = op.load_workbook(file)
    ws   = wb['donnee']

    n        = ws.cell(2, 2).value   # nombre d'objets
    m        = ws.cell(1, 2).value   # nombre de camions (bins)
    capacity = ws.cell(3, 2).value   # capacité

    poids = [ws.cell(5, 2 + j).value for j in range(n)]

    solver = pywraplp.Solver('CuttingStock_MinBobines',
                              pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    # x[i][k] = 1 si objet i assigné au camion k
    x = [[solver.IntVar(0, 1, f'x_{i}_{k}') for k in range(m)]
         for i in range(n)]
    # y[k] = 1 si camion k est utilisé
    y = [solver.IntVar(0, 1, f'y_{k}') for k in range(m)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Chaque objet dans exactement 1 camion
    for i in range(n):
        solver.Add(sum(x[i][k] for k in range(m)) == 1)

    # Capacité
    for k in range(m):
        solver.Add(sum(poids[i] * x[i][k] for i in range(n)) <= capacity * y[k])

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    solver.Minimize(sum(y[k] for k in range(m)))

    start  = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        nb_camions = int(solver.Objective().Value())
        print(f"\nNombre optimal de camions : {nb_camions}")
        print(f"Temps de calcul           : {elapsed:.4f} s")
        print(f"Itérations                : {solver.iterations()}")
        print(f"Nœuds B&B                 : {solver.nodes()}")

        # Écriture résultats
        ws_res = wb['resultat']
        ws_res['A1'] = 'Nombre optimal de camions'
        ws_res['B1'] = nb_camions
        for k in range(m):
            if y[k].solution_value() > 0.5:
                for i in range(n):
                    ws_res.cell(3 + k, 2 + i).value = int(
                        x[i][k].solution_value())
        wb.save(file)
        print(f"Résultats écrits dans     : {file}")
    else:
        print("Pas de solution optimale trouvée.")


# ─────────────────────────────────────────────
# Exercice 3 — Découpe bobines (min perte)
# ─────────────────────────────────────────────
def exercice3_cutting_stock_min_perte():
    print("\n" + "=" * 55)
    print("Exercice 3 — Découpe bobines (min perte)")
    print("=" * 55)

    file = 'documents/tp4exo3.xlsx'
    wb   = op.load_workbook(file)
    ws   = wb['donnee']

    n        = ws.cell(2, 2).value
    m        = ws.cell(1, 2).value
    capacity = ws.cell(3, 2).value

    poids = [ws.cell(5, 2 + j).value for j in range(n)]

    solver = pywraplp.Solver('CuttingStock_MinPerte',
                              pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    x = [[solver.IntVar(0, 1, f'x_{i}_{k}') for k in range(m)]
         for i in range(n)]
    y = [solver.IntVar(0, 1, f'y_{k}') for k in range(m)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    for i in range(n):
        solver.Add(sum(x[i][k] for k in range(m)) == 1)

    for k in range(m):
        solver.Add(sum(poids[i] * x[i][k] for i in range(n)) <= capacity * y[k])

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    # Minimiser la perte totale = capacité utilisée - charge réelle
    perte = sum(
        capacity * y[k] - sum(poids[i] * x[i][k] for i in range(n))
        for k in range(m)
    )
    solver.Minimize(perte)

    start  = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        perte_val  = solver.Objective().Value()
        nb_camions = sum(int(y[k].solution_value()) for k in range(m))
        print(f"\nPerte minimale            : {perte_val:.1f} kg")
        print(f"Nombre de camions utilisés: {nb_camions}")
        print(f"Temps de calcul           : {elapsed:.4f} s")
        print(f"Itérations                : {solver.iterations()}")
        print(f"Nœuds B&B                 : {solver.nodes()}")
    else:
        print("Pas de solution optimale trouvée.")


# ─────────────────────────────────────────────
if __name__ == "__main__":
    exercice1_chargement_avions()
    exercice2_cutting_stock_min_bobines()
    exercice3_cutting_stock_min_perte()