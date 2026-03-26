# ============================================================
# TP2 - Exercices 1 & 2 : Problème de mélange - Fabrication de l'acier
# Recherche Opérationnelle | IMT Mines Alès
# Solveur : OR-Tools (GLOP) via ortools.linear_solver
# ============================================================

from ortools.linear_solver import pywraplp
import openpyxl as op
import time

# ─────────────────────────────────────────────
# Exercice 1 — Modèle explicite (sans Excel)
# ─────────────────────────────────────────────
def exercice1_modele_explicite():
    print("=" * 55)
    print("Exercice 1 — Modèle explicite (sans interface Excel)")
    print("=" * 55)

    solver = pywraplp.Solver('Melange_Acier_Explicite',
                              pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)
    infinity = solver.infinity()

    D = 5000  # Demande totale en kg

    # Variables : quantité de chaque matière première (kg)
    X1 = solver.NumVar(0, 4000,    'Fer1')
    X2 = solver.NumVar(0, 3000,    'Fer2')
    X3 = solver.NumVar(0, 6000,    'Fer3')
    X4 = solver.NumVar(0, 5000,    'Cuivre1')
    X5 = solver.NumVar(0, 2000,    'Cuivre2')
    X6 = solver.NumVar(0, 3000,    'Aluminium1')
    X7 = solver.NumVar(0, 2500,    'Aluminium2')

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Contrainte de demande totale
    solver.Add(X1 + X2 + X3 + X4 + X5 + X6 + X7 == D)

    # Contrainte carbone : 2% <= %C <= 3%
    solver.Add(2.5*X1 + 3*X2 >= 0.02 * D * 100)
    solver.Add(2.5*X1 + 3*X2 <= 0.03 * D * 100)

    # Contrainte cuivre : %Cu <= 0.6%
    solver.Add(0.3*X3 + 90*X4 + 96*X5 + 0.4*X6 + 0.6*X7 <= 0.006 * D * 100)

    # Contrainte manganèse : 1.2% <= %Mn <= 1.65%
    solver.Add(1.3*X1 + 0.8*X2 + 4*X5 + 1.2*X6 <= 1.65 * D * 100)
    solver.Add(1.3*X1 + 0.8*X2 + 4*X5 + 1.2*X6 >= 1.2  * D * 100)

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    # Minimiser le coût
    solver.Minimize(1.20*X1 + 1.50*X2 + 0.90*X3 + 1.30*X4 +
                    1.45*X5 + 1.20*X6 + 1.00*X7)

    start = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        print(f"\nSolution optimale      : {solver.Objective().Value():.2f} €")
        print(f"Temps de calcul        : {elapsed:.4f} s")
        print(f"Itérations             : {solver.iterations()}")
        print("\nQuantités utilisées (kg) :")
        for v in [X1, X2, X3, X4, X5, X6, X7]:
            print(f"  {v.name():12s} : {v.solution_value():.2f} kg")
    else:
        print("Pas de solution optimale trouvée.")


# ─────────────────────────────────────────────
# Exercice 2 — Interface Excel
# ─────────────────────────────────────────────
def exercice2_interface_excel():
    print("\n" + "=" * 55)
    print("Exercice 2 — Interface Excel")
    print("=" * 55)

    file = 'documents/TP2_EX02.xlsx'
    wb = op.load_workbook(file)
    ws = wb['Feuil1']

    # Lecture des données
    matieres  = []
    carbone   = []
    cuivre    = []
    manganese = []
    stock     = []
    cout      = []

    n = 7  # nombre de matières premières
    for i in range(n):
        row = 3 + i  # données à partir de la ligne 3
        matieres.append(ws.cell(row, 1).value)
        carbone.append(ws.cell(row, 2).value   or 0)
        cuivre.append(ws.cell(row, 3).value    or 0)
        manganese.append(ws.cell(row, 4).value or 0)
        stock.append(ws.cell(row, 5).value     or 0)
        cout.append(ws.cell(row, 6).value      or 0)

    D = 5000

    solver = pywraplp.Solver('Melange_Acier_Excel',
                              pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)
    infinity = solver.infinity()

    X = [solver.NumVar(0, stock[i], f'x_{i}') for i in range(n)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Demande totale
    ct_D = solver.Constraint(D, D, 'ct_demande')
    for i in range(n):
        ct_D.SetCoefficient(X[i], 1)

    # Carbone 2% - 3%
    ct_c_min = solver.Constraint(2 * D, infinity,  'ct_carbone_min')
    ct_c_max = solver.Constraint(-infinity, 3 * D, 'ct_carbone_max')
    for i in range(n):
        ct_c_min.SetCoefficient(X[i], carbone[i])
        ct_c_max.SetCoefficient(X[i], carbone[i])

    # Cuivre <= 0.6%
    ct_cu = solver.Constraint(-infinity, 0.6 * D, 'ct_cuivre')
    for i in range(n):
        ct_cu.SetCoefficient(X[i], cuivre[i])

    # Manganèse 1.2% - 1.65%
    ct_mn_min = solver.Constraint(1.2  * D, infinity,  'ct_mn_min')
    ct_mn_max = solver.Constraint(-infinity, 1.65 * D, 'ct_mn_max')
    for i in range(n):
        ct_mn_min.SetCoefficient(X[i], manganese[i])
        ct_mn_max.SetCoefficient(X[i], manganese[i])

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    obj = solver.Objective()
    for i in range(n):
        obj.SetCoefficient(X[i], cout[i])
    obj.SetMinimization()

    start = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        print(f"\nSolution optimale      : {solver.Objective().Value():.2f} €")
        print(f"Temps de calcul        : {elapsed:.4f} s")
        print(f"Itérations             : {solver.iterations()}")

        # Écriture des résultats dans Excel
        ws_out = wb.create_sheet('Résultats')
        ws_out['A1'] = 'Matière première'
        ws_out['B1'] = 'Quantité (kg)'
        ws_out['A2'] = 'Coût total (€)'
        ws_out['B2'] = solver.Objective().Value()
        for i in range(n):
            ws_out.cell(4 + i, 1).value = matieres[i]
            ws_out.cell(4 + i, 2).value = X[i].solution_value()

        wb.save(file)
        print(f"Résultats écrits dans  : {file} (feuille 'Résultats')")

        print("\nQuantités utilisées (kg) :")
        for i in range(n):
            print(f"  {matieres[i]:15s} : {X[i].solution_value():.2f} kg")
    else:
        print("Pas de solution optimale trouvée.")


# ─────────────────────────────────────────────
if __name__ == "__main__":
    exercice1_modele_explicite()
    exercice2_interface_excel()