# ============================================================
# TP2 - Exercice 3 : Planification de production - Aliments bétail
# Recherche Opérationnelle | IMT Mines Alès
# Solveur : OR-Tools (GLOP) via ortools.linear_solver
# ============================================================

from ortools.linear_solver import pywraplp
import openpyxl as op
import time


# ─────────────────────────────────────────────
# Exercice 3 — Planification aliments pour bétail
# ─────────────────────────────────────────────
def exercice3_aliments_betail():
    print("=" * 55)
    print("Exercice 3 — Planification aliments pour bétail")
    print("=" * 55)

    file = 'documents/TP2_EX03.xlsx'
    wb = op.load_workbook(file)
    ws = wb['Feuil1']

    # ── Lecture des données ──
    matieres   = [ws.cell(1, 2+j).value for j in range(3)]   # Avoine, Maïs, Mélasse
    nutriments = [ws.cell(2+i, 1).value for i in range(3)]   # Protéines, Lipides, Glucides

    # Composition nutritionnelle (ligne i, colonne j)
    compo = [[ws.cell(2+i, 2+j).value or 0 for j in range(3)] for i in range(3)]

    # Teneurs souhaitées (colonne 5)
    teneurs    = [ws.cell(2+i, 5).value or 0 for i in range(3)]
    sens       = ['>=', '>=', '<=']  # Protéines >=, Lipides >=, Glucides <=

    # Stocks et prix
    stocks = [ws.cell(6, 2+j).value or 0 for j in range(3)]
    prix   = [ws.cell(7, 2+j).value or 0 for j in range(3)]

    n = 3  # nombre de matières premières

    solver = pywraplp.Solver('Aliments_Betail',
                              pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)
    infinity = solver.infinity()

    X = [solver.NumVar(0, stocks[j], matieres[j]) for j in range(n)]

    print(f"Nombre de variables    : {solver.NumVariables()}")

    # Contraintes nutritionnelles
    D = sum(stocks)  # quantité totale produite = somme des stocks disponibles
    for i in range(3):
        if sens[i] == '>=':
            ct = solver.Constraint(teneurs[i] * D / 100, infinity, f'ct_{nutriments[i]}')
        else:
            ct = solver.Constraint(-infinity, teneurs[i] * D / 100, f'ct_{nutriments[i]}')
        for j in range(n):
            ct.SetCoefficient(X[j], compo[i][j])

    # Contrainte : utiliser tout le stock disponible
    ct_total = solver.Constraint(D, D, 'ct_total')
    for j in range(n):
        ct_total.SetCoefficient(X[j], 1)

    print(f"Nombre de contraintes  : {solver.NumConstraints()}")

    # Minimiser le coût d'achat
    obj = solver.Objective()
    for j in range(n):
        obj.SetCoefficient(X[j], prix[j])
    obj.SetMinimization()

    start  = time.time()
    status = solver.Solve()
    elapsed = time.time() - start

    if status == pywraplp.Solver.OPTIMAL:
        print(f"\nSolution optimale      : {solver.Objective().Value():.2f} €")
        print(f"Temps de calcul        : {elapsed:.4f} s")
        print(f"Itérations             : {solver.iterations()}")

        # Écriture des résultats dans Excel
        try:
            ws_out = wb['Résultats']
        except KeyError:
            ws_out = wb.create_sheet('Résultats')

        ws_out['A1'] = 'Matière première'
        ws_out['B1'] = 'Quantité utilisée (kg)'
        ws_out['A2'] = 'Coût total (€)'
        ws_out['B2'] = round(solver.Objective().Value(), 2)
        for j in range(n):
            ws_out.cell(4 + j, 1).value = matieres[j]
            ws_out.cell(4 + j, 2).value = round(X[j].solution_value(), 2)

        wb.save(file)
        print(f"Résultats écrits dans  : {file} (feuille 'Résultats')")

        print("\nQuantités utilisées (kg) :")
        for j in range(n):
            print(f"  {matieres[j]:12s} : {X[j].solution_value():.2f} kg")
    else:
        print("Pas de solution optimale trouvée.")



# ─────────────────────────────────────────────
if __name__ == "__main__":
    exercice3_aliments_betail()